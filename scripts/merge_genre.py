#!/usr/bin/env python3
"""把 data/genre/ 下多份跨源 genre 映射表合并为单一权威表 genre.csv。

设计要点
--------
- 合并键：以原始标签（日文 `ja` 优先）去重。
  * jav321 把日文放在 `translate` 列（`ja` 多为空），自动把该列提升为合并键。
  * 无日文来源（javdb）按 `en` / `zh_cn` / `id` 兜底作为合并键。
- 值列补全（让单一文档本身就是完整可读的资料库）：
  * `zh_cn`：优先 data/zh.json 的 tag_zh[ja] 人工精修（最高优先，可纠正源站错误）
    -> 否则 source zh_cn -> 否则 源 translate（若为中文）繁->简 ->
    否则 源 zh_tw（若为中文）繁->简 -> 否则保留日文原文（避免中文视图丢标签）。
    （注：个别源站把日文原词抄进 zh_tw 列，故 translate 优先于 zh_tw 且仅采纳含中文的候选；
    人工精修层优先级高于一切源站，与 genre_norm 运行时行为一致。）
  * `zh_tw`：优先 source zh_tw -> 否则 zhconv(源 zh_cn 简->繁)，两列互补填满。
- 去重：同一合并键只保留一行，多来源用 `source` 列记录溯源。
- 输出整洁、可排序、UTF-8-BOM（Excel 直接正确打开中文）。

输出
----
- data/genre/genre.csv  ：结构化的单一权威映射表（供 scripts/genre_norm.py 消费，也是可读资料库）
- data/genre/genre.xlsx ：带样式的可读版本（冻结表头 + 彩色表头 + 自动列宽 + 筛选器）

用法
----
    python scripts/merge_genre.py
"""
import csv
import json
import os
import sys
import glob
import zhconv

_HERE = os.path.dirname(os.path.abspath(__file__))
_GENRE_DIR = os.path.join(_HERE, "..", "data", "genre")
_OUT_CSV = os.path.join(_GENRE_DIR, "genre.csv")
_OUT_XLSX = os.path.join(_GENRE_DIR, "genre.xlsx")

# 源优先级：值列取「首个非空」时据此排序（越靠前越优先）
_PRIORITY = ["javdb", "javbus", "javlib", "avsox", "jav321"]

# 输出列（与 genre_norm 的键/值列保持一致，便于直接消费）
FIELDS = ["id", "url", "ja", "zh_cn", "zh_tw", "en", "translate", "note", "source"]


def _src_name(fn):
    base = os.path.basename(fn)
    if base.startswith("genre_") and base.endswith(".csv"):
        return base[len("genre_"):-len(".csv")]
    return base[:-4]


def _load_sources():
    """读取原始跨源 CSV。优先读 legacy/ 归档，否则读 data/genre/ 下除 genre.csv 外的文件。"""
    legacy = os.path.join(_GENRE_DIR, "legacy")
    if os.path.isdir(legacy):
        pat = os.path.join(legacy, "genre_*.csv")
    else:
        pat = os.path.join(_GENRE_DIR, "genre*.csv")
    files = [f for f in glob.glob(pat)
             if os.path.basename(f) != "genre.csv"]
    return sorted(files)


def _ja_eff(row, src):
    """解析该行的「原始标签键」：优先 ja；jav321 的日文在 translate 列；否则空。"""
    ja = (row.get("ja") or "").strip()
    if ja:
        return ja
    if src == "jav321":
        tr = (row.get("translate") or "").strip()
        if tr:
            return tr  # jav321 把日文放在 translate 列
    return ""


def _is_chinese(s):
    """含汉字且不含假名，视为中文（可用于补全 zh_cn）。"""
    has_han = False
    for ch in s:
        o = ord(ch)
        if 0x4E00 <= o <= 0x9FFF or 0x3400 <= o <= 0x4DBF:
            has_han = True
        elif 0x3040 <= o <= 0x30FF:  # 假名 -> 日文，排除
            return False
    return has_han


def _fill_zh_cn(members, ja, zh_json):
    """按优先级补全简中：zh.json 人工精修 -> 源 zh_cn -> 源 translate(中,繁->简) -> 源 zh_tw(中,繁->简) -> 原文。

    注意：
    - data/zh.json 的 tag_zh 是人工精修层，优先级最高（可纠正源站错误，如把
      源站误写的「和服・丧服」纠正为「和服・浴衣」），且需与 genre_norm 运行时一致
      （运行时 zh.json 最后应用、覆盖 CSV）。
    - zh_tw 列在个别源站里被直接填了日文（如 アスリート 的 zh_tw 也是 アスリート），
      因此先于 zh_tw 检查 translate，且仅在「该候选确实含中文」时才采用，
      避免把日文原词当中文填进 zh_cn。
    """
    if ja and ja in zh_json:
        return zh_json[ja]
    for _, row, _ in members:
        v = (row.get("zh_cn") or "").strip()
        if v:
            return v
    for _, row, _ in members:
        v = (row.get("translate") or "").strip()
        if v and _is_chinese(v):
            return zhconv.convert(v, "zh-cn")
    for _, row, _ in members:
        v = (row.get("zh_tw") or "").strip()
        if v and _is_chinese(v):
            return zhconv.convert(v, "zh-cn")
    return ja  # 无可译内容时保留原文，避免中文视图丢标签


def _fill_zh_tw(members):
    """补全繁中：源 zh_tw -> 源 zh_cn(简->繁)。"""
    for _, row, _ in members:
        v = (row.get("zh_tw") or "").strip()
        if v:
            return v
    for _, row, _ in members:
        v = (row.get("zh_cn") or "").strip()
        if v:
            return zhconv.convert(v, "zh-tw")
    return ""


def merge():
    # 载入人工精修的中文映射（data/zh.json 的 tag_zh），作为 zh_cn 的最高优先级补充
    zh_json = {}
    try:
        with open(os.path.join(_HERE, "..", "data", "zh.json"), encoding="utf-8") as f:
            zh_json = (json.load(f).get("tag_zh") or {})
    except Exception:
        pass

    groups = {}  # key -> list[(priority_idx, row, src)]
    for fp in _load_sources():
        src = _src_name(fp)
        try:
            with open(fp, encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    key = _ja_eff(row, src)
                    if not key:
                        key = ((row.get("en") or "").strip()
                               or (row.get("zh_cn") or "").strip()
                               or (row.get("id") or "").strip())
                    if not key:
                        continue
                    pidx = _PRIORITY.index(src) if src in _PRIORITY else 99
                    groups.setdefault(key, []).append((pidx, row, src))
        except Exception as e:  # 容错：单文件损坏不影响整体
            print("跳过 %s: %s" % (fp, e), file=sys.stderr)

    rows = []
    empty_before = 0
    for key, members in groups.items():
        members.sort(key=lambda m: m[0])
        merged = {fld: "" for fld in FIELDS}
        merged["ja"] = key  # 原始标签键（日文优先，缺则英文/中文/站点id）
        merged["zh_cn"] = _fill_zh_cn(members, key, zh_json)
        if not merged["zh_cn"]:
            empty_before += 1
        merged["zh_tw"] = _fill_zh_tw(members)
        merged["en"] = _pick(members, "en")
        merged["translate"] = _pick(members, "translate")
        merged["id"] = (members[0][1].get("id") or "").strip()
        merged["url"] = (members[0][1].get("url") or "").strip()
        notes = []
        for _, row, _ in members:
            n = (row.get("note") or "").strip()
            if n and n not in notes:
                notes.append(n)
        merged["note"] = " / ".join(notes)
        merged["source"] = ";".join(sorted({m[2] for m in members}))
        rows.append(merged)

    # 排序：日文/原文在前，便于浏览
    rows.sort(key=lambda r: (r["ja"] or "\uffff", r["zh_cn"] or ""))

    with open(_OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print("合并完成：%d 个唯一标签 -> %s（补全后空 zh_cn: %d）"
          % (len(rows), _OUT_CSV, empty_before))
    return rows


def _pick(members, valcol):
    """按源优先级取首个非空值列。"""
    for _, row, _ in members:
        v = (row.get(valcol) or "").strip()
        if v:
            return v
    return ""


def write_xlsx(rows):
    """生成带样式的可读 xlsx（CSV 无法承载样式，xlsx 可以）。

    样式：
    - 粉色加粗表头 + 全边框 + 冻结首行 + 自动筛选
    - 斑马纹（隔行浅粉）提升可读性
    - 中文友好字体（微软雅黑），垂直居中、左对齐
    - 「简中仍为日文」的行整行琥珀色高亮，提醒需人工补译
    - 额外「说明」工作表解释各列含义与维护方式
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print("跳过 xlsx（openpyxl 不可用）: %s" % e, file=sys.stderr)
        return

    # 仅含「人读资料库」需要的列（id/url 内部字段不展示）
    disp = [
        ("ja", "原始标签 (ja)"),
        ("zh_cn", "简中 (zh_cn)"),
        ("zh_tw", "繁中 (zh_tw)"),
        ("en", "英文 (en)"),
        ("translate", "源中文 (translate)"),
        ("note", "说明 (note)"),
        ("source", "来源 (source)"),
    ]
    headers = [h for _, h in disp]
    ncol = len(headers)

    # 判断「简中仍为日文」：含平/片假名即视为未译（排除中点「・」与重复记号「ヽヾ」，
    # 否则像「和服・浴衣」这类纯中文、仅用「・」分隔的词会被误判）
    def _is_untranslated(r):
        cn = (r.get("zh_cn") or "").strip()
        return any("\u3040" <= ch <= "\u30FF" and ch not in "・ヽヾ" for ch in cn)

    # ---- 样式常量 ----
    HEAD_FILL = PatternFill("solid", fgColor="FF5C8A")
    HEAD_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="微软雅黑", size=10.5)
    ZEBRA = PatternFill("solid", fgColor="FCEFF4")          # 隔行浅粉
    NEEDED = PatternFill("solid", fgColor="FFF2CC")          # 未译行：琥珀色
    thin = Side(style="thin", color="E6D2DA")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "genre"
    ws.append(headers)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    n_untrans = 0
    for ridx, r in enumerate(rows, start=2):
        vals = [r.get(col, "") for col, _ in disp]
        ws.append(vals)
        untr = _is_untranslated(r)
        if untr:
            n_untrans += 1
            fill = NEEDED
        elif ridx % 2 == 0:
            fill = ZEBRA
        else:
            fill = None
        for c in range(1, ncol + 1):
            cell = ws.cell(row=ridx, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (3, 4, 7) else LEFT
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(ncol), len(rows) + 1)
    ws.sheet_view.showGridLines = False
    for i, width in enumerate([34, 24, 22, 26, 22, 30, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 26

    # ---- 说明工作表 ----
    sheet2 = wb.create_sheet("说明")
    sheet2.sheet_view.showGridLines = False
    lines = [
        ("jav-idol-db · 标签资料库 (genre)", True),
        ("", False),
        ("本表是唯一的标签权威文件，同时承担两项职责：", False),
        ("  1) 站点构建 / 搜索检索的来源（scripts/genre_norm.py 只读取本表）", False),
        ("  2) 你可在 Excel 中直接阅读、筛选、补译的资料库", False),
        ("", False),
        ("各列含义", True),
        ("  原始标签 (ja)   —— 日文 / 英文原始标签，是合并去重的主键", False),
        ("  简中 (zh_cn)    —— 简体中文译名（站点中文视图显示这一列）", False),
        ("  繁中 (zh_tw)    —— 繁体中文译名（由简中互补生成，供繁中用户）", False),
        ("  英文 (en)       —— 英文译名", False),
        ("  源中文 (translate) —— 各源站（javbus/javlib 等）提供的原站翻译；", False),
        ("                     构建时作为备选补全『简中』，平时可忽略（非『别名』）", False),
        ("  说明 (note)     —— 备注信息", False),
        ("  来源 (source)   —— 该标签由哪些源站合并而来（如 javbus;javlib）", False),
        ("", False),
        ("琥珀色整行 = 简中仍为日文", True),
        ("  表示暂无可考的中文译名，多为品牌 / 奖项专名（如 AV OPEN 赛事分区）。", False),
        ("  若你确知其中文，可直接在 genre.csv 的『简中』列填写后，重跑合并脚本。", False),
        ("", False),
        ("如何重新生成", True),
        ("  修改 data/genre/genre.csv（或 legacy/ 下源文件）后，在仓库根目录运行：", False),
        ("      python scripts/merge_genre.py", False),
        ("  会刷新 genre.csv 与 genre.xlsx，并请用 build_index.py 重建站点数据。", False),
    ]
    sheet2.column_dimensions["A"].width = 92
    for i, (text, bold) in enumerate(lines, start=1):
        c = sheet2.cell(row=i, column=1, value=text)
        c.font = Font(name="微软雅黑", bold=bold, size=12 if bold and i == 1 else 10.5,
                      color="FF5C8A" if (bold and i == 1) else "333333")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb.save(_OUT_XLSX)
    print("已生成样式化 xlsx: %s（未译行高亮 %d 行）" % (_OUT_XLSX, n_untrans))


if __name__ == "__main__":
    rows = merge()
    write_xlsx(rows)
