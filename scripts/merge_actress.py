#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把分散在四处的女优信息聚合为单一可读权威表 actress.csv + 样式化 actress.xlsx。

背景
----
标签侧已有 data/genre/genre.csv（给人读的资料库 + 站点检索来源）。
女优侧此前没有对应物：档案散在 data/actresses/<名>/profile.json，
作品归属散在 data/works/*.json，中文名在 data/zh.json，别名在 data/actress/alias.json。
本脚本把这些信息聚成一张表，让人能像查标签表一样直接查女优。

数据源（均为「手写源」；本表是派生的只读视图）
------------------------------------------
- data/actresses/<名>/profile.json   档案：生日 / 身高 / 三围 / 罩杯 / 出生地 / 血型 / 事务所 / 简介 / 头像
- data/works/*.json                  按归属女优统计：作品数 / 首作日期 / 最新作品日期
- data/zh.json 的 actress_zh         中文名（与站点中文视图同一显示层，优先级最高）
- data/actress/alias.json            别名簇（经 scripts/alias_norm.py「整组 cluster 并入」展开）

铁律（与 genre 表一致的架构约定）
--------------------------------
本表是**生成物**，不要手改 —— 直接改 actress.csv 会在下次重跑时被覆盖丢失。
要改内容请改上面的源，然后重跑本脚本：改档案 -> profile.json；改中文名 -> zh.json；
改别名 -> alias.json。改完再跑 build_index.py 让站点同步。

输出
----
- data/actresses/actress.csv   结构化单一权威表（UTF-8-BOM，Excel 直接打开不乱码）
- data/actresses/actress.xlsx  带样式的可读版（冻结表头 + 彩色表头 + 斑马纹 + 自动筛选 + 说明页）

用法
----
    python scripts/merge_actress.py
"""
import csv
import glob
import json
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
_BASE = os.path.join(_HERE, "..")
_ACT_DIR = os.path.join(_BASE, "data", "actresses")
_OUT_CSV = os.path.join(_ACT_DIR, "actress.csv")
_OUT_XLSX = os.path.join(_ACT_DIR, "actress.xlsx")

# 复用站点同一套别名归一逻辑，保证表里的名字与站点显示完全一致
sys.path.insert(0, _HERE)
try:
    from alias_norm import normalize_actress, actress_search_terms
except Exception:  # pragma: no cover - 极端缺依赖时不阻断，退化为原名
    normalize_actress = lambda n: n
    actress_search_terms = lambda n: [n]
try:
    from actress_status import status_label
except Exception:  # pragma: no cover
    def status_label(code, lang="zh"):
        return code or ""

# 非个人女优的特殊归属（合集厂牌 / 无主聚合桶），alias_norm 里同样被排除出精选集合
_SPECIAL = {
    "S1オールスター": "合集厂牌（非个人女优），无独立档案",
    "其他作品": "无主作品聚合桶（非真实女优）",
}

# 无女优归属的作品归入此聚合桶（与站点 build_index.py 的 OTHER 口径保持一致）
_OTHER = "其他作品"

# CSV 列（内部字段名）—— 新增 status 系列（在役/引退/出道/引退/复出）
FIELDS = [
    "name", "name_zh", "type", "status", "status_zh", "aliases",
    "work_count", "first_work", "latest_work",
    "debut_year", "debut_date", "retire_date", "comeback_date", "status_source",
    "birthdate", "height", "cup", "measurements",
    "birthplace", "blood_type", "agency",
    "bio", "avatar", "source", "updated_at", "note",
]


def _load_profiles():
    """读取全部女优档案：{女优名: profile 字典}。"""
    out = {}
    for fp in sorted(glob.glob(os.path.join(_ACT_DIR, "*", "profile.json"))):
        # 目录名即女优名（档案里的 name 字段可能与目录名不一致，以目录名为准）
        name = os.path.basename(os.path.dirname(fp))
        try:
            with open(fp, encoding="utf-8") as f:
                out[name] = json.load(f)
        except Exception as e:
            print("跳过档案 %s: %s" % (fp, e), file=sys.stderr)
    return out


def _load_name_zh():
    """读取中文名映射 data/zh.json 的 actress_zh（与站点同一显示层）。"""
    try:
        with open(os.path.join(_BASE, "data", "zh.json"), encoding="utf-8") as f:
            return (json.load(f).get("actress_zh") or {})
    except Exception:
        return {}


def _load_work_stats():
    """扫描作品库，按「归一后的女优名」统计 作品数 / 首作日期 / 最新作品日期。

    归一用 alias_norm.normalize_actress，这样「田中レモン」这类别名会被并到
    「楓カレン」名下，统计口径与站点分组完全一致。
    """
    cnt = {}
    dates = {}
    for fp in glob.glob(os.path.join(_BASE, "data", "works", "*.json")):
        try:
            with open(fp, encoding="utf-8") as f:
                w = json.load(f)
        except Exception:
            continue
        raw = w.get("actress") or ((w.get("actresses") or [None])[0])
        if not raw:
            # 无女优归属的作品归入「其他作品」聚合桶，与站点 build_index 的 OTHER 口径一致。
            # 保留它才能让表的作品总数与作品库对得上账（否则会凭空少掉一批）。
            raw = _OTHER
        name = normalize_actress(raw)
        cnt[name] = cnt.get(name, 0) + 1
        d = (w.get("date") or "").strip()
        if d:
            dates.setdefault(name, []).append(d)
    stats = {}
    for name, n in cnt.items():
        ds = sorted(dates.get(name, []))
        stats[name] = {
            "work_count": n,
            "first_work": ds[0] if ds else "",
            "latest_work": ds[-1] if ds else "",
        }
    return stats


def _aliases_of(name):
    """返回该女优的别名列表（不含本人），来自 alias_norm 的整组 cluster。"""
    try:
        terms = actress_search_terms(name) or []
    except Exception:
        terms = []
    # 去掉与本人同名的项，其余按原序保留
    return [t for t in terms if t and t != name]


def _s(v):
    """把任意值安全转为去空格字符串（None / 空 -> ''）。"""
    if v is None:
        return ""
    if isinstance(v, (list, tuple)):
        return ";".join(str(x) for x in v if x)
    return str(v).strip()


def merge():
    profiles = _load_profiles()
    name_zh = _load_name_zh()
    stats = _load_work_stats()

    # 女优全集 = 有档案的 ∪ 有作品的（保证不漏任何一位，哪怕只有作品没有档案）
    names = set(profiles) | set(stats)

    rows = []
    for name in names:
        p = profiles.get(name, {})
        st = stats.get(name, {})
        note = _SPECIAL.get(name, "")
        r = {f: "" for f in FIELDS}
        r["name"] = name
        r["name_zh"] = name_zh.get(name, "")
        r["type"] = "女优" if not note else "合集厂牌" if "合集" in note else "聚合桶"
        r["status"] = _s(p.get("status")) or "unknown"
        r["status_zh"] = status_label(p.get("status")) or "不明"
        r["aliases"] = ";".join(_aliases_of(name))
        r["work_count"] = st.get("work_count", 0)
        r["first_work"] = st.get("first_work", "")
        r["latest_work"] = st.get("latest_work", "")
        # 档案字段：缺档案时整列留空，表里一眼能看出「这位还没建档案」
        for k in ("debut_year", "debut_date", "retire_date", "comeback_date",
                  "status_source", "birthdate", "height", "cup", "measurements",
                  "birthplace", "blood_type", "agency", "bio", "avatar",
                  "source", "updated_at"):
            r[k] = _s(p.get(k))
        # 档案里自带的 aliases 也并进别名列（去重、保持顺序）
        if p.get("aliases"):
            extra = [a for a in p["aliases"]
                     if a and a != name and a not in r["aliases"].split(";")]
            if extra:
                r["aliases"] = ";".join([x for x in r["aliases"].split(";") if x] + extra)
        if not note and not p:
            note = "有作品但尚无 profile.json 档案"
        r["note"] = note
        rows.append(r)

    # 排序：真实女优在前、作品多的更前（符合浏览直觉）；合集厂牌 / 聚合桶沉到底部
    rows.sort(key=lambda r: (r["type"] != "女优", -int(r["work_count"] or 0), r["name"]))

    with open(_OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    n_noprofile = sum(1 for r in rows if not r["source"])
    print("聚合完成：%d 位 -> %s（缺档案 %d 位）" % (len(rows), _OUT_CSV, n_noprofile))
    return rows


def write_xlsx(rows):
    """生成带样式的可读 xlsx（CSV 承载不了样式，xlsx 可以）。

    样式沿用 genre.xlsx 的规格，保持两套资料库观感一致：
    - 粉色加粗表头 + 全边框 + 冻结首行 + 自动筛选
    - 斑马纹（隔行浅粉）
    - 「非个人女优 / 缺档案」的行整行琥珀高亮，提示需要补档案
    - 头像列做成可点击超链接，直接看图
    - 额外「说明」工作表解释各列与维护方式
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print("跳过 xlsx（openpyxl 不可用）: %s" % e, file=sys.stderr)
        return

    # (内部字段, 表头) —— 只展示「人读资料库」需要的列
    disp = [
        ("name", "女优名 (日文)"),
        ("name_zh", "中文名"),
        ("type", "类型"),
        ("status", "状态"),
        ("status_zh", "状态(中)"),
        ("aliases", "别名"),
        ("work_count", "作品数"),
        ("first_work", "首作"),
        ("latest_work", "最新作"),
        ("debut_year", "出道年"),
        ("debut_date", "出道日期"),
        ("retire_date", "引退日期"),
        ("comeback_date", "复出日期"),
        ("status_source", "状态来源"),
        ("birthdate", "生日"),
        ("height", "身高(cm)"),
        ("cup", "罩杯"),
        ("measurements", "三围"),
        ("birthplace", "出生地"),
        ("blood_type", "血型"),
        ("agency", "事务所"),
        ("bio", "简介"),
        ("avatar", "头像"),
        ("source", "数据来源"),
        ("updated_at", "更新"),
        ("note", "备注"),
    ]
    headers = [h for _, h in disp]
    ncol = len(headers)

    HEAD_FILL = PatternFill("solid", fgColor="FF5C8A")
    HEAD_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="微软雅黑", size=10.5)
    LINK_FONT = Font(name="微软雅黑", size=10.5, color="185FA5", underline="single")
    ZEBRA = PatternFill("solid", fgColor="FCEFF4")
    NEEDED = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="E6D2DA")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "actress"
    ws.append(headers)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    n_flag = 0
    for ridx, r in enumerate(rows, start=2):
        # 需要关注 = 非个人女优（合集厂牌/聚合桶）/完全没有档案/状态未知需核查
        need = ((r.get("type") != "女优") or (not r.get("source"))
                or (r.get("type") == "女优" and r.get("status") == "unknown"))
        if need:
            n_flag += 1
            fill = NEEDED
        elif ridx % 2 == 0:
            fill = ZEBRA
        else:
            fill = None
        for c, (fld, _h) in enumerate(disp, start=1):
            v = r.get(fld, "")
            # 作品数/身高/出道年写成数值，便于排序筛选
            if fld in ("work_count", "height", "debut_year"):
                try:
                    v = int(v) if str(v).strip() else ""
                except Exception:
                    pass
            cell = ws.cell(row=ridx, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (
                3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 23, 24, 25
            ) else LEFT
            if fill:
                cell.fill = fill
            # 头像列：有链接就做成可点击超链接，显示为「查看」
            if fld == "avatar" and v:
                cell.hyperlink = str(v)
                cell.value = "查看"
                cell.font = LINK_FONT

    ws.freeze_panes = "B2"  # 冻结表头 + 首列（女优名），横向滚动时不丢主键
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(ncol), len(rows) + 1)
    ws.sheet_view.showGridLines = False
    for i, width in enumerate(
            [16, 14, 9, 9, 9, 30, 8, 12, 12, 9, 12, 12, 12, 13, 12, 9, 7, 16,
             12, 7, 18, 46, 8, 12, 12, 26],
            start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 26

    # ---- 说明工作表 ----
    ws2 = wb.create_sheet("说明")
    ws2.sheet_view.showGridLines = False
    total_works = sum(int(r.get("work_count") or 0) for r in rows)
    lines = [
        ("jav-idol-db · 女优资料库 (actress)", True),
        ("", False),
        ("共 %d 位，合计 %d 部作品。按作品数从多到少排列。" % (len(rows), total_works), False),
        ("", False),
        ("各列含义", True),
        ("  女优名 (日文) —— 主键，日文原名；作品归属已按别名归一（如 田中レモン 并入 楓カレン）", False),
        ("  中文名       —— data/zh.json 的 actress_zh，站点中文视图显示这一列", False),
        ("  类型         —— 女优 / 合集厂牌 / 聚合桶（后两者不是真实个人）", False),
        ("  别名         —— 同一人的其他艺名（简繁中、旧艺名），来自 alias.json 整组 cluster", False),
        ("  作品数 / 首作 / 最新作 —— 由 data/works 实时统计，口径与站点分组一致", False),
        ("  出道年 ~ 事务所 —— 来自 data/actresses/<名>/profile.json 的档案字段", False),
        ("  状态 / 状态(中) —— 在役 / 引退 / 休业，来自 profile.status（抓取来源见「状态来源」列）", False),
        ("  出道日期 / 引退日期 / 复出日期 —— 来自 profile 的 debut_date / retire_date / comeback_date", False),
        ("  状态来源      —— wikipedia-ja / avjoho / researched 等，标记状态信息出处，便于人工复核", False),
        ("  头像         —— 点击「查看」直接打开图片链接", False),
        ("  数据来源      —— researched 表示人工核查过的档案；空表示尚未建档", False),
        ("", False),
        ("琥珀色整行 = 需要关注", True),
        ("  非个人女优（合集厂牌 / 聚合桶），或有作品但还没建 profile.json 档案。", False),
        ("", False),
        ("重要：本表是生成物，请勿直接手改", True),
        ("  actress.csv / actress.xlsx 由脚本生成，手改会在重跑时被覆盖。", False),
        ("  要改内容请改「源」，再重跑脚本：", False),
        ("    改档案字段 -> data/actresses/<名>/profile.json", False),
        ("    改中文名   -> data/zh.json 的 actress_zh", False),
        ("    改别名     -> data/actress/alias.json", False),
        ("  然后在仓库根目录运行：", False),
        ("      python scripts/merge_actress.py", False),
        ("  最后用 build_index.py 重建站点数据。", False),
    ]
    ws2.column_dimensions["A"].width = 96
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws2.cell(row=i, column=1, value=text)
        c.font = Font(name="微软雅黑", bold=bold, size=12 if bold and i == 1 else 10.5,
                      color="FF5C8A" if (bold and i == 1) else "333333")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb.save(_OUT_XLSX)
    print("已生成样式化 xlsx: %s（需关注行高亮 %d 行）" % (_OUT_XLSX, n_flag))


if __name__ == "__main__":
    write_xlsx(merge())
